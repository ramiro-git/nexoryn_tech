import csv
import io
import datetime
from typing import List, Dict, Any, Optional
from fpdf import FPDF
import openpyxl

class ExportService:
    @staticmethod
    def export_to_csv(data: List[Dict[str, Any]], filename: str = "export.csv") -> str:
        if not data:
            return ""
        
        # Determine headers from the first item keys
        headers = list(data[0].keys())
        
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=headers)
        writer.writeheader()
        
        # Format rows before writing
        formatted_data = []
        for row in data:
            formatted_data.append({k: ExportService._format_value(v) for k, v in row.items()})
            
        writer.writerows(formatted_data)
        
        return output.getvalue()

    @staticmethod
    def export_to_excel(data: List[Dict[str, Any]], filename: str = "export.xlsx") -> str:
        if not data:
            return ""
            
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"
        
        headers = list(data[0].keys())
        ws.append(headers)
        
        for row in data:
            ws.append([ExportService._format_value(row.get(h)) for h in headers])
            
        # Save to a temporary buffer or file? 
        # For simplicity in desktop app, we might return bytes or save to a known path.
        # Here we will assume the caller handles file writing if we return bytes, 
        # BUT GenericTable logic usually picks a path. 
        # Let's adjust: Return bytes content usually best for flexibility.
        
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue() # Returns bytes

    @staticmethod
    def _sanitize(text: str) -> str:
        """Sanitize text for FPDF/Latin-1 compatibility."""
        if not text:
            return ""
        # Replace common incompatible characters
        replacements = {
            "—": "-",  # em-dash
            "–": "-",  # en-dash
            "“": '"',
            "”": '"',
            "‘": "'",
            "’": "'",
            "…": "...",
        }
        for k, v in replacements.items():
            text = text.replace(k, v)
        
        # Enforce latin-1, replacing unknown chars with ?
        try:
            return text.encode("latin-1", "replace").decode("latin-1")
        except:
            return text

    @staticmethod
    def _format_value(val: Any) -> str:
        """Format values for export (Booleans to Spanish, None/Placeholders to empty)."""
        if val is None or val == "" or val == "—" or val == "--":
            return ""
        if isinstance(val, bool):
            return "Sí" if val else "No"
        return str(val)

    @staticmethod
    def export_to_pdf(data: List[Dict[str, Any]], title: str = "Export") -> bytes:
        if not data:
            return b""

        headers = list(data[0].keys())
        col_count = len(headers)

        # Use Landscape. Switch to A3 if too many columns
        fmt = "A3" if col_count > 12 else "A4"
        pdf = FPDF(orientation="L", format=fmt)
        pdf.add_page()
        
        # Dynamic font size based on column count
        base_font_size = 10
        if col_count > 15:
            base_font_size = 7
        elif col_count > 10:
            base_font_size = 8
        
        pdf.set_font("helvetica", size=base_font_size)
        
        # Sanitize Title
        safe_title = ExportService._sanitize(title)
        
        # Title
        pdf.set_font("helvetica", style="B", size=16)
        pdf.cell(0, 10, txt=safe_title, ln=True, align="C")
        pdf.ln(5)
        
        # Timestamp
        pdf.set_font("helvetica", size=8)
        pdf.cell(0, 5, txt=f"Generado: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M')}", ln=True, align="R")
        pdf.ln(5)
        
        # Calculate available width
        page_width = pdf.w - 2 * pdf.l_margin
        
        # Smart column width calculation
        # 1. Measure max content width for each column (sample first 50 rows)
        col_widths_map = {h: pdf.get_string_width(str(h).upper()) + 4 for h in headers} # +4 for padding
        
        for row in data[:50]:
            for h in headers:
                val = ExportService._format_value(row.get(h))
                safe_val = ExportService._sanitize(val)
                w = pdf.get_string_width(safe_val) + 4
                if w > col_widths_map[h]:
                    col_widths_map[h] = w
        
        # 2. Normalize to fit page width
        total_content_width = sum(col_widths_map.values())
        final_col_widths = []
        
        if total_content_width > page_width:
            # Scale down proportionally if too wide
            scale_factor = page_width / total_content_width
            for h in headers:
                final_col_widths.append(col_widths_map[h] * scale_factor)
        else:
            # Distribute extra space evenly
            extra_space = page_width - total_content_width
            extra_per_col = extra_space / col_count
            for h in headers:
                final_col_widths.append(col_widths_map[h] + extra_per_col)
        
        # Header
        pdf.set_font("helvetica", style="B", size=base_font_size)
        for i, h in enumerate(headers):
            pdf.cell(final_col_widths[i], 8, txt=ExportService._sanitize(str(h).upper()), border=1, align="C")
        pdf.ln()
        
        # Rows
        pdf.set_font("helvetica", size=base_font_size)
        for row in data:
            # Check page break
            if pdf.get_y() > (pdf.h - 20):
                pdf.add_page()
                # Repeat header
                pdf.set_font("helvetica", style="B", size=base_font_size)
                for i, h in enumerate(headers):
                    pdf.cell(final_col_widths[i], 8, txt=ExportService._sanitize(str(h).upper()), border=1, align="C")
                pdf.ln()
                pdf.set_font("helvetica", size=base_font_size)

            max_line_height = 8
            # In complex tables we might need multi-cell, but for now stick to single cell clipping
            # to keep alignment simple.
            for i, h in enumerate(headers):
                val = ExportService._format_value(row.get(h))
                safe_val = ExportService._sanitize(val)
                
                # Check if fits, truncate if scaled down too much
                # (Simple truncation based on char count is risky with proportional fonts, 
                # but simplistic check: text width vs col width)
                
                cell_w = final_col_widths[i]
                text_w = pdf.get_string_width(safe_val)
                
                display_val = safe_val
                if text_w > (cell_w - 2):
                    # Rough truncation
                    while pdf.get_string_width(display_val + "...") > (cell_w - 2) and len(display_val) > 0:
                        display_val = display_val[:-1]
                    display_val += "..."
                    
                pdf.cell(cell_w, max_line_height, txt=display_val, border=1)
            pdf.ln()
            
        return pdf.output()  # Returns bytes
